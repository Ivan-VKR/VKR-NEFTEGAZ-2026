"""Модуль сохранения результатов в TXT и XLSX."""

from __future__ import annotations

from pathlib import Path
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from calculations import CalculationResult


HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def save_txt_report(file_path: str, report_text: str) -> None:
    Path(file_path).write_text(report_text, encoding="utf-8")


def save_xlsx_report(
    file_path: str,
    package: Dict[str, object],
    report_text: str,
) -> None:
    calc: CalculationResult = package["calculation_result"]
    constructive_solution = package["constructive_solution"]
    coatings = package["coatings"]
    equipment = package.get("equipment", {"items": [], "normative_notes": []})

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Итоги"

    ws1["A1"] = "Расчетно-рекомендательный отчет"
    ws1["A1"].font = Font(bold=True, size=14)

    row = 3
    ws1[f"A{row}"] = "Параметр"
    ws1[f"B{row}"] = "Значение"
    for cell in (ws1[f"A{row}"], ws1[f"B{row}"]):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    for key, value in calc.summary.items():
        ws1[f"A{row}"] = key
        ws1[f"B{row}"] = value
        row += 1

    row += 1
    ws1[f"A{row}"] = "Рекомендуемое конструктивное решение"
    ws1[f"A{row}"].font = BOLD_FONT
    row += 1
    ws1[f"A{row}"] = "Вариант"
    ws1[f"B{row}"] = constructive_solution["solution"]
    row += 1
    ws1[f"A{row}"] = "Обоснование"
    ws1[f"B{row}"] = constructive_solution["justification"]
    ws1[f"B{row}"].alignment = WRAP_ALIGNMENT
    row += 1
    ws1[f"A{row}"] = "Баллы выбора"
    ws1[f"B{row}"] = (
        f"продукт={constructive_solution.get('product_score', '-')}; "
        f"температура={constructive_solution.get('temperature_score', '-')}; "
        f"объем={constructive_solution.get('volume_score', '-')}; "
        f"итого={constructive_solution.get('total_score', '-')}"
    )
    row += 2

    ws1[f"A{row}"] = "Примечание"
    ws1[f"A{row}"].font = BOLD_FONT
    row += 1
    ws1[f"A{row}"] = (
        "Предварительный модуль для ВКР; для рабочего проектирования необходима полная нормативная проверка."
    )
    ws1[f"A{row}"].alignment = WRAP_ALIGNMENT

    ws1.column_dimensions["A"].width = 42
    ws1.column_dimensions["B"].width = 95
    for row_cells in ws1.iter_rows(min_row=1, max_row=row, min_col=1, max_col=2):
        for cell in row_cells:
            cell.alignment = WRAP_ALIGNMENT

    ws2 = wb.create_sheet("Пояса стенки")
    headers = [
        "Пояс",
        "Уровень от низа, м",
        "Давление, МПа",
        "Требуемая толщина, мм",
        "Принятая толщина, мм",
        "Окружное напряжение, МПа",
        "Проверка",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, course in enumerate(calc.course_results, start=2):
        ws2.cell(row=row_idx, column=1, value=course.course_no)
        ws2.cell(row=row_idx, column=2, value=course.level_from_bottom_m)
        ws2.cell(row=row_idx, column=3, value=course.pressure_mpa)
        ws2.cell(row=row_idx, column=4, value=course.required_thickness_mm)
        ws2.cell(row=row_idx, column=5, value=course.adopted_thickness_mm)
        ws2.cell(row=row_idx, column=6, value=course.hoop_stress_mpa)
        ws2.cell(row=row_idx, column=7, value="OK" if course.strength_ok else "НЕ OK")

    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws2.column_dimensions[col_letter].width = 20

    ws3 = wb.create_sheet("Варианты проката")
    steel_headers = [
        "Класс прочности",
        "Пример марки",
        "σдоп, МПа",
        "Мин. расчетная температура, °C",
        "Лист стенки",
        "Листов на пояс",
        "Всего листов стенки",
        "Диапазон толщин стенки, мм",
        "Лист днища",
        "Листов днища",
        "Лист крыши",
        "Листов крыши",
        "Масса, т",
        "Оценка применимости",
        "Примечание",
    ]
    for col, header in enumerate(steel_headers, start=1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, variant in enumerate(calc.steel_variants, start=2):
        ws3.cell(row=row_idx, column=1, value=variant.strength_class)
        ws3.cell(row=row_idx, column=2, value=variant.example_grade)
        ws3.cell(row=row_idx, column=3, value=variant.allowable_stress_mpa)
        ws3.cell(row=row_idx, column=4, value=variant.min_design_temp_c)
        ws3.cell(row=row_idx, column=5, value=variant.shell_sheet_format)
        ws3.cell(row=row_idx, column=6, value=variant.shell_sheets_per_course)
        ws3.cell(row=row_idx, column=7, value=variant.shell_total_sheets)
        ws3.cell(row=row_idx, column=8, value=variant.shell_thickness_range_mm)
        ws3.cell(row=row_idx, column=9, value=variant.bottom_sheet_format)
        ws3.cell(row=row_idx, column=10, value=variant.bottom_sheet_count)
        ws3.cell(row=row_idx, column=11, value=variant.roof_sheet_format)
        ws3.cell(row=row_idx, column=12, value=variant.roof_sheet_count)
        ws3.cell(row=row_idx, column=13, value=variant.estimated_total_mass_t)
        ws3.cell(row=row_idx, column=14, value=variant.suitability)
        ws3.cell(row=row_idx, column=15, value=variant.note)

    for col_letter, width in {
        "A": 16, "B": 26, "C": 12, "D": 18, "E": 16, "F": 14, "G": 18,
        "H": 24, "I": 16, "J": 12, "K": 16, "L": 12, "M": 12, "N": 28, "O": 52,
    }.items():
        ws3.column_dimensions[col_letter].width = width
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT

    ws4 = wb.create_sheet("Покрытия")
    coating_headers = ["Зона", "Система", "Толщина", "Слои", "Область применения", "Срок службы"]
    for col, header in enumerate(coating_headers, start=1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, (_, zone_data) in enumerate(coatings.items(), start=2):
        ws4.cell(row=row_idx, column=1, value=zone_data["zone"])
        ws4.cell(row=row_idx, column=2, value=zone_data["system"])
        ws4.cell(row=row_idx, column=3, value=zone_data["thickness"])
        ws4.cell(row=row_idx, column=4, value=zone_data["layers"])
        ws4.cell(row=row_idx, column=5, value=zone_data["application"])
        ws4.cell(row=row_idx, column=6, value=zone_data["service_life"])

    for col_letter, width in {"A": 28, "B": 46, "C": 18, "D": 12, "E": 45, "F": 18}.items():
        ws4.column_dimensions[col_letter].width = width
    for row in ws4.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT

    ws5 = wb.create_sheet("Оборудование")
    equipment_headers = ["Наименование", "Количество", "Проход DN", "Назначение", "Обоснование"]
    for col, header in enumerate(equipment_headers, start=1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    row_idx = 2
    for item in equipment.get("items", []):
        ws5.cell(row=row_idx, column=1, value=item["name"])
        ws5.cell(row=row_idx, column=2, value=item["qty"])
        ws5.cell(row=row_idx, column=3, value=item.get("diameter_dn", "-"))
        ws5.cell(row=row_idx, column=4, value=item["purpose"])
        ws5.cell(row=row_idx, column=5, value=item["basis"])
        row_idx += 1
    if equipment.get("normative_notes"):
        row_idx += 1
        ws5.cell(row=row_idx, column=1, value="Нормативные примечания")
        ws5.cell(row=row_idx, column=1).font = BOLD_FONT
        row_idx += 1
        for note in equipment["normative_notes"]:
            ws5.cell(row=row_idx, column=1, value=note)
            row_idx += 1

    for col_letter, width in {"A": 34, "B": 16, "C": 14, "D": 40, "E": 60}.items():
        ws5.column_dimensions[col_letter].width = width
    for row in ws5.iter_rows(min_row=2, max_col=5):
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT

    ws6 = wb.create_sheet("Текст отчета")
    ws6["A1"] = report_text
    ws6["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws6.column_dimensions["A"].width = 120

    wb.save(file_path)
